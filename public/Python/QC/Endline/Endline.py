# # # import os, sys
# # # from pathlib import Path
# # # import numpy as np
# # # import mysql.connector
# # # from datetime import datetime
# # # from sqlalchemy import create_engine
# # # import pandas as pd

# # # hostname='pbvweb01v'
# # # engine = create_engine('mysql+mysqlconnector://root:123456@pbvweb01v:3306/pr2k', echo=False)
# # # def get_this_week():
# # #     week_str=''
# # #     week=1+int(datetime.now().strftime("%W"))
# # #     if week<10:
# # #         week_str='W0'+str(week)
# # #     else:
# # #         week_str='W'+str(week)
# # #     return week_str

# # # def get_date_format(date):
# # #     year=date[0:4]
# # #     month=date[5:7]
# # #     day=date[8:10]
# # #     # return year+month+day
# # #     return year+month+day, '_'+day+'-'+month+'-'+year+'_'

# # # today=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
# # # date, dateQuery=get_date_format(today)

# # # #=====Daily Bundle==========================
# # # if __name__=="__main__":
# # #     uptoHours=today[11:13]
# # #     uptoHours_int=int(uptoHours)#-1
# # #     # uptoHours=str(uptoHours_int)
# # #     # if uptoHours_int<10:
# # #     #     uptoHours='0'+uptoHours
# # #     fromHours_int=uptoHours_int-1
# # #     fromHours=str(fromHours_int)
# # #     today_date=today[0:10]
# # #     if fromHours_int<10:
# # #         fromHours='0'+fromHours
# # #     if uptoHours_int>=6 and uptoHours_int<=24:
# # #         dailyBundle_link='\\\\pbvfps1\\PBShare2\\Scan\\Report\\QCReconcile\\'
# # #         thisWeek=get_this_week()
# # #         if not os.path.exists(dailyBundle_link+thisWeek):
# # #             os.makedirs(dailyBundle_link+thisWeek)
# # #         if not os.path.exists(dailyBundle_link+thisWeek+'\\'+date):
# # #             os.makedirs(dailyBundle_link+thisWeek+'\\'+date)
# # #         dailyBundle=pd.read_sql('select TICKET, EMPLOYEE, DATE, QC, IRR, IRR1, IRR2, MODIFIED, FILE, TimeUpdate, TimeModified from employee_scanticket where DATE="'+date+'" and TimeUpdate>="'+today_date+' '+fromHours+':30:00" and TimeUpdate<="'+today_date+' '+uptoHours+':30:00";', engine)
# # #         # dailyBundle=pd.read_sql('select TICKET, EMPLOYEE, DATE, QC, IRR, IRR1, IRR2, MODIFIED, FILE, TimeUpdate, TimeModified from employee_scanticket where FILE LIKE "%'+dateQuery+fromHours+'%" OR FILE LIKE "%'+dateQuery+uptoHours+'%";', engine)
# # #         engine.dispose()
# # #         try:
# # #             dailyBundle.to_excel(dailyBundle_link+thisWeek+'\\'+date+'\\QCReconcile_'+uptoHours+'.xlsx', index=False)
# # #         except:
# # #             print('Cant save QCReconcile_'+uptoHours+'.xlsx')
# # #         else:
# # #             print('QCReconcile_'+uptoHours+'.xlsx is Created')
# # #     else:
# # #         print('Out of report time range')
# # # -*- coding: utf-8 -*-
# # """
# # Created on Thu Jul  2 17:22:56 2020

# # @author: dule4
# # """


# # import numpy as np
# # import mysql.connector, os, sys
# # from datetime import datetime
# # from sqlalchemy import create_engine
# # import pandas as pd

# # hostname='pbvweb01v'
# # engine = create_engine('mysql+mysqlconnector://root:123456@pbvweb01v:3306/pr2k', echo=False)
# # engineNam = create_engine('mysql+mysqlconnector://root:123456@pbvweb01v:3306/erpsystem', echo=False)
# # def get_this_week():
# #     week_str=''
# #     week=1+int(datetime.now().strftime("%W"))
# #     if week<10:
# #         week_str='W0'+str(week)
# #     else:
# #         week_str='W'+str(week)
# #     return week_str

# # def get_date_format(date):
# #     year=date[0:4]
# #     month=date[5:7]
# #     day=date[8:10]
# #     # return year+month+day
# #     return year+month+day, '_'+day+'-'+month+'-'+year+'_', year+'-'+month+'-'+day



# # def getShift(thisWeek, thisHour):
# #     thisWeek_int=int(thisWeek[1:3])
# #     shift=""
# #     if thisWeek_int%2!=0:
# #         if int(thisHour)<14:
# #             shift="R"
# #         else:
# #             shift="B"
# #     else:
# #         if int(thisHour)<14:
# #             shift="B"
# #         else:
# #             shift="R"
# #     return shift

# # #=====Daily Bundle==========================
# # if __name__=="__main__":
# #     today=sys.argv[1]
# #     timeSpan=sys.argv[2]
# #     # datetime.now().strftime('%Y-%m-%d %H:%M:%S')
# #     date, dateQuery, dateQ=get_date_format(today)
# #     uptoHours=today[11:13]
# #     uptoHours_int=int(uptoHours)-1
# #     uptoHours=str(uptoHours_int)
# #     today_date=today[0:10]
# #     if uptoHours_int<10:
# #         uptoHours='0'+uptoHours
# #     fromHours_int=uptoHours_int-1
# #     fromHours=str(fromHours_int)
# #     today_date=today[0:10]
    
# #     if fromHours_int<10:
# #         fromHours='0'+fromHours

# #     if timeSpan=='1h':
# #         fromHours=uptoHours
# #         fromHours_int=uptoHours_int

# #     if uptoHours_int>=6 and uptoHours_int<=24:
# #         dailyBundle_link='D:\\HanesApp\\public\\\Python\\QC\\Endline\\Report\\'
# #         thisWeek=get_this_week()
# #         # if not os.path.exists(dailyBundle_link+thisWeek):
# #         #     os.makedirs(dailyBundle_link+thisWeek)
# #         # if not os.path.exists(dailyBundle_link+thisWeek+'\\'+date):
# #         #     os.makedirs(dailyBundle_link+thisWeek+'\\'+date)
# #         sql=('SELECT LEFT(FILE, 6) AS LINE, MID(FILE, 7,1) AS SHIFT, DATE, COUNT(DISTINCT BUNDLE) AS SAMPLE FROM employee_scanticket '
# #              +' WHERE QC!="" AND QC!="000000" AND (FILE LIKE "%'+dateQuery+fromHours+'%" OR FILE LIKE "%'+dateQuery+uptoHours+'%")'
# #              +' GROUP BY LEFT(FILE, 6);')
# #         # sql=('SELECT LEFT(FILE, 6) AS LINE, MID(FILE, 7,1) AS SHIFT, DATE, COUNT(DISTINCT BUNDLE) AS SAMPLE FROM employee_scanticket '
# #         #      +' WHERE QC!="" AND QC!="000000" and File like "%'+dateQuery+' '+fromHours+':00:00" and TimeUpdate<="'+dateQ+' '+uptoHours+':00:00" '
# #         #      +' GROUP BY LEFT(FILE, 6);')
# #         # print(sql)
# #         data=pd.read_sql(sql, engine)
# #         engine.dispose()
# #         data_json=[]
# #         irr_json=[]
# #         shift_real=getShift(thisWeek, uptoHours)
# #         for row in range(0, len(data)):
# #             line=data.iloc[row, 0]
# #             shift=data.iloc[row, 1]
# # #            date=data.iloc[row, 2]
# #             sample=data.iloc[row, 3]*2
# #             sql=('SELECT TICKET, EMPLOYEE, QC, IRR, IRR1, IRR2, FILE FROM employee_scanticket '
# #              +' WHERE QC!="" AND QC!="000000" AND (FILE LIKE "'+line+shift+dateQuery+fromHours+'%" OR FILE LIKE "'+line+shift+dateQuery+uptoHours+'%") and (IRR!="000" OR IRR1!="000" OR IRR2!="000");')
# #             data_irr=pd.read_sql(sql, engine)
# #             sum_irr=0
# #             # print(sql)
# #             for row_irr in range(0, len(data_irr)):
# #                 ticket=data_irr.iloc[row_irr, 0]
# #                 employee=data_irr.iloc[row_irr, 1]
# #                 namepd=pd.read_sql('select Name, Shift from setup_emplist where ID like "%'+employee+'";', engineNam)
# #                 engineNam.dispose()
# #                 name=''
# #                 if len(namepd)>0:
# #                     name=namepd.iloc[0,0]
# #                     shift_emp=namepd.iloc[0,1]
# #                 qc=data_irr.iloc[row_irr, 2]
# #                 file=data_irr.iloc[row_irr, 6]
# #                 irr=data_irr.iloc[row_irr, 3]
# #                 if irr!='000':
# #                     if shift_real[0]==shift_emp[0]:
# #                         sum_irr=sum_irr+1
# #                     irr_json.append({'Group':line, 'Shift': shift_real, 'Ticket':ticket, 'Employee':employee, 'Name':name, 'Shift_emplist':shift_emp, 'QC':qc,'IRR':irr, 'FILE':file})
# #                 irr1=data_irr.iloc[row_irr, 4]
# #                 if irr1!='000':
# #                     if shift_real[0]==shift_emp[0]:
# #                         sum_irr=sum_irr+1
# #                     irr_json.append({'Group':line, 'Shift': shift_real, 'Ticket':ticket, 'Employee':employee, 'Name':name, 'Shift_emplist':shift_emp, 'QC':qc,'IRR':irr1, 'FILE':file})
# #                 irr2=data_irr.iloc[row_irr, 5]
# #                 if irr2!='000':
# #                     if shift_real[0]==shift_emp[0]:
# #                         sum_irr=sum_irr+1
# #                     irr_json.append({'Group':line, 'Shift': shift_real, 'Ticket':ticket, 'Employee':employee, 'Name':name, 'Shift_emplist':shift_emp, 'QC':qc,'IRR':irr2, 'FILE':file})
# #             endline=0
# #             if (line=='200234'):
# #                 sample=sample/2*36
# #                 endline=round(sum_irr*100/sample,2)
# #             else:
# #                 sample=sample+sum_irr*2
# #                 endline=round(sum_irr*100/sample,2)
# #             data_json.append({'Group':line, 'Shift': shift_real, 'Date':date, 'FromTime':fromHours, 'ToTime':uptoHours, 'Sample':sample, 'Defect': sum_irr, 'EndLine':endline})
# #         dataF=pd.DataFrame(data_json)
# #         irrF=pd.DataFrame(irr_json)
# # #        dailyBundle=pd.read_sql('select TICKET, EMPLOYEE, DATE, QC, IRR, IRR1, IRR2, MODIFIED, FILE, TimeUpdate, TimeModified from employee_scanticket where DATE="'+date+'" and TimeUpdate>="'+today_date+' '+fromHours+':30:00" and TimeUpdate<="'+today_date+' '+uptoHours+':30:00";', engine)
# #         dailyBundle=pd.read_sql('select TICKET, EMPLOYEE, DATE, QC, IRR, IRR1, IRR2, MODIFIED, FILE, TimeUpdate, TimeModified from employee_scanticket where FILE LIKE "%'+dateQuery+fromHours+'%" OR FILE LIKE "%'+dateQuery+uptoHours+'%";', engine)
# #         engine.dispose()
# #         link='QCReconcile_'+uptoHours+'.xlsx'
# #         try:
# #             writer = pd.ExcelWriter(dailyBundle_link+link, engine='xlsxwriter')
# #             dailyBundle.to_excel(writer,sheet_name='Raw Data', index=False)
# #             dataF.to_excel(writer,sheet_name='Summary Endline', index=False)
# #             irrF.to_excel(writer,sheet_name='Detail Defect', index=False)
# #             writer.save()
# #         except:
# #         #     # print('Cant save QCReconcile_'+uptoHours+'.xlsx')
# #             print('fail')
# #         else:
# #             print(link)
# #     else:
# #         # print('Out of report time range')
# #         a=1
            
        
# """
# Created on Thu Jul  2 17:22:56 2020

# @author: dule4
# """
# """

# import numpy as np
# import mysql.connector, os
# from datetime import datetime
# from sqlalchemy import create_engine
# import pandas as pd

# hostname='pbvweb01v'
# engine = create_engine('mysql+mysqlconnector://root:123456@pbvweb01v:3306/pr2k', echo=False)
# engineNam = create_engine('mysql+mysqlconnector://root:123456@pbvweb01v:3306/erpsystem', echo=False)
# def get_this_week():
#     week_str=''
#     week=1+int(datetime.now().strftime("%W"))
#     if week<10:
#         week_str='W0'+str(week)
#     else:
#         week_str='W'+str(week)
#     return week_str

# def get_date_format(date):
#     year=date[0:4]
#     month=date[5:7]
#     day=date[8:10]
#     # return year+month+day
#     return year+month+day, '_'+day+'-'+month+'-'+year+'_', year+'-'+month+'-'+day

# today=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
# date, dateQuery, dateQ=get_date_format(today)

# def getShift(thisWeek, thisHour):
#     thisWeek_int=int(thisWeek[1:3])
#     shift=""
#     if thisWeek_int%2!=0:
#         if int(thisHour)<14:
#             shift="R"
#         else:
#             shift="B"
#     else:
#         if int(thisHour)<14:
#             shift="B"
#         else:
#             shift="R"
#     return shift

# #=====Daily Bundle==========================
# if __name__=="__main__":
#     h=today[11:13]
#     uptoHours=today[11:13]
#     uptoHours_int=int(uptoHours)-1
#     uptoHours=str(uptoHours_int)
#     today_date=today[0:10]
#     if uptoHours_int<10:
#         uptoHours='0'+uptoHours
#     fromHours_int=uptoHours_int-1
#     fromHours=str(fromHours_int)
#     today_date=today[0:10]
#     if fromHours_int<10:
#         fromHours='0'+fromHours
#     if uptoHours_int>=6 and uptoHours_int<=24:
#         dailyBundle_link='\\\\pbvfps1\\PBShare2\\Scan\\Report\\QCReconcile\\'
#         thisWeek=get_this_week()
#         if not os.path.exists(dailyBundle_link+thisWeek):
#             os.makedirs(dailyBundle_link+thisWeek)
#         if not os.path.exists(dailyBundle_link+thisWeek+'\\'+date):
#             os.makedirs(dailyBundle_link+thisWeek+'\\'+date)
#         sql=('SELECT LEFT(FILE, 6) AS LINE, MID(FILE, 7,1) AS SHIFT, DATE, COUNT(DISTINCT BUNDLE) AS SAMPLE FROM employee_scanticket '
#             + ' WHERE QC!="" AND QC!="000000" AND (FILE LIKE "%'+dateQuery+fromHours+'%" OR FILE LIKE "%'+dateQuery+uptoHours+'%")'
#             + ' GROUP BY LEFT(FILE, 6);')
#         # sql=('SELECT LEFT(FILE, 6) AS LINE, MID(FILE, 7,1) AS SHIFT, DATE, COUNT(DISTINCT BUNDLE) AS SAMPLE FROM employee_scanticket '
#         #      +' WHERE QC!="" AND QC!="000000" and File like "%'+dateQuery+' '+fromHours+':00:00" and TimeUpdate<="'+dateQ+' '+uptoHours+':00:00" '
#         #      +' GROUP BY LEFT(FILE, 6);')
#         # print(sql)
#         data=pd.read_sql(sql, engine)
#         engine.dispose()
#         data_json=[]
#         irr_json=[]
#         other_json=[]
#         shift_real=getShift(thisWeek, uptoHours)
#         for row in range(0, len(data)):
#             line=data.iloc[row, 0]
#             shift=data.iloc[row, 1]
# #            date=data.iloc[row, 2]
#             sample=data.iloc[row, 3]*2
#             sql=('SELECT TICKET, EMPLOYEE, QC, IRR, IRR1, IRR2, FILE FROM employee_scanticket '
#              +' WHERE QC!="" AND QC!="000000" AND (FILE LIKE "'+line+shift+dateQuery+fromHours+'%" OR FILE LIKE "'+line+shift+dateQuery+uptoHours+'%") and (IRR!="000" OR IRR1!="000" OR IRR2!="000");')
#             data_irr=pd.read_sql(sql, engine)
#             sum_irr=0
#             # print(sql)
#             for row_irr in range(0, len(data_irr)):
#                 ticket=data_irr.iloc[row_irr, 0]
#                 employee=data_irr.iloc[row_irr, 1]
#                 namepd=pd.read_sql('select Name, Shift from setup_emplist where ID like "%'+employee+'";', engineNam)
#                 engineNam.dispose()
#                 name=''
#                 if len(namepd)>0:
#                     name=namepd.iloc[0,0]
#                     shift_emp=namepd.iloc[0,1]
#                 qc=data_irr.iloc[row_irr, 2]
#                 file=data_irr.iloc[row_irr, 6]
#                 irr=data_irr.iloc[row_irr, 3]
#                 if irr!='000' and int(irr)<30:
#                     if shift_real[0]==shift_emp[0]:
#                         sum_irr=sum_irr+1
#                         irr_json.append({'Group':line, 'Shift': shift_real, 'Ticket':ticket, 'Employee':employee, 'Name':name, 'Shift_emplist':shift_emp, 'QC':qc,'IRR':irr, 'FILE':file})
#                     else:
#                         other_json.append({'Group':line, 'Shift': shift_real, 'Ticket':ticket, 'Employee':employee, 'Name':name, 'Shift_emplist':shift_emp, 'QC':qc,'IRR':irr, 'FILE':file})
#                 irr1=data_irr.iloc[row_irr, 4]

#                 if irr1!='000' and int(irr)<30:
#                     if shift_real[0]==shift_emp[0]:
#                         sum_irr=sum_irr+1
#                         irr_json.append({'Group':line, 'Shift': shift_real, 'Ticket':ticket, 'Employee':employee, 'Name':name, 'Shift_emplist':shift_emp, 'QC':qc,'IRR':irr1, 'FILE':file})
#                     else:
#                         other_json.append({'Group':line, 'Shift': shift_real, 'Ticket':ticket, 'Employee':employee, 'Name':name, 'Shift_emplist':shift_emp, 'QC':qc,'IRR':irr, 'FILE':file})
#                 irr2=data_irr.iloc[row_irr, 5]
#                 if irr2!='000' and int(irr)<30:
#                     if shift_real[0]==shift_emp[0]:
#                         sum_irr=sum_irr+1
#                         irr_json.append({'Group':line, 'Shift': shift_real, 'Ticket':ticket, 'Employee':employee, 'Name':name, 'Shift_emplist':shift_emp, 'QC':qc,'IRR':irr2, 'FILE':file})
#                     else:
#                         other_json.append({'Group':line, 'Shift': shift_real, 'Ticket':ticket, 'Employee':employee, 'Name':name, 'Shift_emplist':shift_emp, 'QC':qc,'IRR':irr, 'FILE':file})
#             endline=0
#             if (line=='200234'):
#                 sample=sample/2*36
#                 endline=round(sum_irr*100/sample,2)
#             else:
#                 sample=sample+sum_irr*2
#                 endline=round(sum_irr*100/sample,2)
#             data_json.append({'Group':line, 'Shift': shift_real, 'Date':date, 'FromTime':fromHours, 'ToTime':uptoHours, 'Sample':sample, 'Defect': sum_irr, 'EndLine':endline})
#         dataF=pd.DataFrame(data_json)
#         irrF=pd.DataFrame(irr_json)
#         otherF=pd.DataFrame(other_json)
# #        dailyBundle=pd.read_sql('select TICKET, EMPLOYEE, DATE, QC, IRR, IRR1, IRR2, MODIFIED, FILE, TimeUpdate, TimeModified from employee_scanticket where DATE="'+date+'" and TimeUpdate>="'+today_date+' '+fromHours+':30:00" and TimeUpdate<="'+today_date+' '+uptoHours+':30:00";', engine)
#         dailyBundle=pd.read_sql('select TICKET, EMPLOYEE, DATE, QC, IRR, IRR1, IRR2, MODIFIED, FILE, TimeUpdate, TimeModified from employee_scanticket where FILE LIKE "%'+dateQuery+fromHours+'%" OR FILE LIKE "%'+dateQuery+uptoHours+'%";', engine)
#         engine.dispose()
#         try:
#             writer = pd.ExcelWriter(dailyBundle_link+thisWeek+'\\'+date+'\\QCReconcile_'+h+'.xlsx', engine='xlsxwriter')
#             dailyBundle.to_excel(writer, sheet_name='Raw Data'       , index=False)
#             dataF.to_excel      (writer, sheet_name='Summary Endline', index=False)
#             irrF.to_excel       (writer, sheet_name='Detail Defect'  , index=False)
#             otherF.to_excel     (writer, sheet_name='Other Shift'    , index=False)
#             writer.save()
#         except:
#         #     # print('Cant save QCReconcile_'+uptoHours+'.xlsx')
#             print('fail;-;-;-')
#         else:
#             print(date+'-'+h+'h')
#     else:
#         # print('Out of report time range')
#         a=1
            
# """
# # -*- coding: utf-8 -*-
# """
# Created on Tue Aug 11 15:49:32 2020

# @author: dule4
# """

# import numpy as np
# import mysql.connector, os
# from datetime import datetime, timedelta
# from sqlalchemy import create_engine
# import pandas as pd
# import openpyxl

# hostname='pbvweb01v'
# engine = create_engine('mysql+mysqlconnector://root:123456@pbvweb01v:3306/pr2k', echo=False)

# def get_this_week():
#     week_str=''
#     week=1+int(datetime.now().strftime("%W"))
#     if week<10:
#         week_str='W0'+str(week)
#     else:
#         week_str='W'+str(week)
#     return week_str

# def get_date_format(date):
#     year=date[0:4]
#     month=date[5:7]
#     day=date[8:10]
#     # today_format = datetime(int(year), int(month), int(day))
#     # yeste_format_datetime =today_format-timedelta(days=1)
#     # yesterday=yeste_format_datetime.strf
#     # return year+month+day
#     return year+month+day, '_'+day+'-'+month+'-'+year+'_', year+'-'+month+'-'+day

# today         = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
# yesterday     = datetime.now()-timedelta(days=1)
# yesterday_str = yesterday.strftime("%Y-%m-%d")
# date, dateQuery, dateQ = get_date_format(today)
# yest, yestQuery, yestQ = get_date_format(yesterday_str)

# def getShift(thisWeek, thisHour):
    
#     thisWeek_int=int(thisWeek[1:3])
#     shift=""
#     if thisWeek_int%2!=0:
#         if int(thisHour)<14.5:
#             shift="R"
#         else:
#             shift="B"
#     else:
#         if int(thisHour)<14.5:
#             shift="B"
#         else:
#             shift="R"
#     return shift

# def getShiftByDate(date, thisHour):
#     shift_query=pd.read_sql('select StartTime, FinishTime, Shift from operation_schedule where DATE="'+date+'"', engine)
#     engine.dispose()
#     shift=''
#     if len(shift_query)>0:
#         start=shift_query.iloc[0, 0]
#         finis=shift_query.iloc[0, 1]
#         shift_first=shift_query.iloc[0, 2]
#         if thisHour<=finis and thisHour>=start:
#             shift=shift_first
#         else:
#             if shift_first=='R':
#                 shift='B'
#             else:
#                 shift='R'
#     return shift

# #=====Daily Bundle==========================
# if __name__=="__main__":
#     h=today[11:13]
#     uptoHours=today[11:13]
#     uptoHours_int=int(uptoHours)-1
#     uptoHours=str(uptoHours_int)
#     today_date=today[0:10]
#     if uptoHours_int<10:
#         uptoHours='0'+uptoHours
#     fromHours_int=uptoHours_int-1
#     fromHours=str(fromHours_int)
#     today_date=today[0:10]
#     if fromHours_int<10:
#         fromHours='0'+fromHours
#     try:
#     # if uptoHours_int>=6 and uptoHours_int<=24:
#         dailyBundle_link='\\\\pbvfps1\\PBShare2\\Scan\\Report\\QCReconcile\\'
#         thisWeek=get_this_week()
#         if not os.path.exists(dailyBundle_link+thisWeek):
#             os.makedirs(dailyBundle_link+thisWeek)
#         if not os.path.exists(dailyBundle_link+thisWeek+'\\'+date):
#             os.makedirs(dailyBundle_link+thisWeek+'\\'+date)
#         sql=("SELECT Ticket, scan.EMPLOYEE as ID_5, ID as ID_6, NAME, LINE, LEFT(Shift,1) as SHIFT, OPERATION, scan.WORK_LOT, active.LOCATION as GROUP_WL, SELLING_STYLE, "
#             +" scan.SIZE, scan.COLOR, IRR, CONCAT(LEFT(FILE, 3),'-',MID(FILE,4,3)) AS GROUP_SCAN, MID(FILE,9, 10) as DATE_SCAN, MID(FILE, 20, 2) as HOUR_SCAN, FILE, QC "
#             +" FROM (SELECT qc.TICKET, scan.EMPLOYEE, qc.IRR, scan.WORK_LOT, qc.FILE, scan.OPERATION, scan.SIZE, scan.COLOR, scan.QC FROM employee_scanticket scan "
#             +" INNER JOIN qc_endline_record qc on scan.TICKET=qc.TICKET ) as scan LEFT JOIN worklot_active active ON scan.WORK_LOT=active.WORK_LOT "
#             +" left join (select ID, RIGHT(ID, 5) as EMPLOYEE, Name, Line, Shift from erpsystem.setup_emplist) as Temp1 on Temp1.EMPLOYEE=scan.EMPLOYEE "
#             +" WHERE (FILE like '%R_18-08-2020%' or FILE like '%B_19-08-2020%') AND IRR!='000' order by DATE_SCAN, HOUR_SCAN, GROUP_WL;")
#         data=pd.read_sql(sql, engine)
#         data_json=[]
#         wb = openpyxl.load_workbook(r'D:\HanesApp\public\Python\QC\Endline\Template\EndlineTemplate.xlsx')#dailyOutput_link+thisWeek+'\\'+date+'\\'+today_date+'_'+thisHour+'h50.xlsx')  
#         row_in_shift=1
#         # row_out_shift=1
#         shift_real=getShift(thisWeek, uptoHours)
#         for row in range(0, len(data)):
#             ticket=str(data.iloc[row, 0])
#             ID=data.iloc[row, 2]
#             name=data.iloc[row, 3]
#             line=data.iloc[row, 4]
#             shift=data.iloc[row, 5]
#             op=data.iloc[row, 6]
#             wl=data.iloc[row, 7]
#             wl_group=data.iloc[row, 8]
#             selling=data.iloc[row, 9]
#             size=str(data.iloc[row, 10])
#             if size=='None':
#                 size=''
#             color=str(data.iloc[row, 11])
#             if color=='None':
#                 color=''
#             irr=data.iloc[row, 12]
#             scan_group=data.iloc[row, 13]
#             scan_day=data.iloc[row, 14]
#             scan_hour=data.iloc[row, 15]
#             file=data.iloc[row, 16]
#             qc=data.iloc[row, 17]
#             if shift==shift_real or (shift!=shift_real and scan_day=='19-08-2020'):
#                 wb.active=0
#                 sheet = wb.active
#                 row_in_shift+=1
#                 sheet.cell(row=row_in_shift, column=1).value=ticket
#                 sheet.cell(row=row_in_shift, column=2).value=ID
#                 sheet.cell(row=row_in_shift, column=3).value=name
#                 sheet.cell(row=row_in_shift, column=4).value=line
#                 sheet.cell(row=row_in_shift, column=5).value=shift_real
#                 sheet.cell(row=row_in_shift, column=6).value=shift
#                 sheet.cell(row=row_in_shift, column=7).value=op
#                 sheet.cell(row=row_in_shift, column=8).value=wl
#                 sheet.cell(row=row_in_shift, column=9).value=wl_group
#                 sheet.cell(row=row_in_shift, column=10).value=selling
#                 sheet.cell(row=row_in_shift, column=11).value=size
#                 sheet.cell(row=row_in_shift, column=12).value=color
#                 sheet.cell(row=row_in_shift, column=13).value=irr
#                 sheet.cell(row=row_in_shift, column=14).value=scan_group
#                 sheet.cell(row=row_in_shift, column=15).value=scan_day
#                 sheet.cell(row=row_in_shift, column=16).value=scan_hour
#                 sheet.cell(row=row_in_shift, column=17).value=str(file)
#                 sheet.cell(row=row_in_shift, column=18).value=qc
#         wb.active=1
#         sheet = wb.active
#         group_list=data['GROUP_SCAN'].unique()
#         group_list.sort()
#         for i in range(0, len(group_list)):
#             group=group_list[i]
#             group_query=data.query('GROUP_SCAN=="'+group+'"')
#             de_lai=0
#             trong_ca=0
#             ngoai_ca=0
#             for j in range(0, len(group_query)):
#                 file=group_query.iloc[j, 16]
#                 shift_emp =group_query.iloc[j, 5]
#                 shift_scan=group_query.iloc[j, 16][6:7]
#                 # shift_real=getShift(thisWeek, uptoHours)
#                 if shift_emp==shift_real and shift_scan==shift_real:
#                     trong_ca+=1
#                 elif shift_emp!=shift_real and shift_scan==shift_real:
#                     ngoai_ca+=1
#                 elif shift_emp==shift_real and shift_scan!=shift_real:
#                     de_lai+=1
#             total=trong_ca+ngoai_ca+de_lai
#             sheet.cell(row=i+3, column=9).value=trong_ca
#             sheet.cell(row=i+3, column=1).value=group
#             sheet.cell(row=i+3, column=2).value=shift_real
#             sheet.cell(row=i+3, column=3).value=de_lai
#             if shift_real=='R':
#                 sheet.cell(row=i+3, column=4).value=trong_ca
#                 sheet.cell(row=i+3, column=5).value=ngoai_ca
#             else:
#                 sheet.cell(row=i+3, column=5).value=trong_ca
#                 sheet.cell(row=i+3, column=4).value=ngoai_ca
#             sheet.cell(row=i+3, column=6).value=total
#         wb.active=2
#         sheet = wb.active
#         query=("select CONCAT(LEFT(FILE, 3),'-',MID(FILE,4,3)) AS GROUP_SCAN, COUNT(DISTINCT BUNDLE) as SAMPLE "
#                 +"from employee_scanticket where FILE like '%B_19-08-2020%' and QC!='000000' and QC!='' and QC!='999999' GROUP BY LEFT(FILE, 6);")
#         data_sample=pd.read_sql(query, engine)
#         if len(data_sample)>0:
#             for i in range(0, len(data_sample)):
#                 group=str(data_sample.iloc[i, 0])
#                 sample=data_sample.iloc[i, 1]
#                 sheet.cell(row=i+2, column=1).value=group
#                 sheet.cell(row=i+2, column=2).value=sample
#         link=''
#         try:
#             link=dailyBundle_link+thisWeek+'\\'+date+'\\QCReconcile_'+uptoHours+'.xlsx'
#             wb.save(dailyBundle_link+thisWeek+'\\'+date+'\\QCReconcile_'+uptoHours+'.xlsx') 
#         except:
#             link=dailyBundle_link+thisWeek+'\\'+date+'\\QCReconcile_'+uptoHours+'_web.xlsx'
#             wb.save(dailyBundle_link+thisWeek+'\\'+date+'\\QCReconcile_'+uptoHours+'_web.xlsx') 
#         print(link)
#     except:
#         print('fail')
            
# -*- coding: utf-8 -*-
"""
Created on Tue Aug 11 15:49:32 2020

@author: dule4
"""

import numpy as np
import mysql.connector, os
from datetime import datetime, timedelta
from sqlalchemy import create_engine
import pandas as pd
import openpyxl, sys

hostname='pbvweb01v'
engine = create_engine('mysql+mysqlconnector://root:123456@pbvweb01v:3306/pr2k', echo=False)

def get_this_week():
    week_str=''
    week=1+int(datetime.now().strftime("%W"))
    if week<10:
        week_str='W0'+str(week)
    else:
        week_str='W'+str(week)
    return week_str

def get_date_format(date):
    year=date[0:4]
    month=date[5:7]
    day=date[8:10]
    return year+month+day, '_'+day+'-'+month+'-'+year+'_', day+'-'+month+'-'+year

def getShift(thisWeek, thisHour):
    
    thisWeek_int=int(thisWeek[1:3])
    shift=""
    if thisWeek_int%2!=0:
        if int(thisHour)<14.5:
            shift="R"
        else:
            shift="B"
    else:
        if int(thisHour)<14.5:
            shift="B"
        else:
            shift="R"
    return shift

def getShiftByDate(date, thisHour):
    shift_query=pd.read_sql('select StartTime, FinishTime, Shift, Note from operation_schedule where DATE="'+date+'"', engine)
    engine.dispose()
    shift=''
    if len(shift_query)>0:
        start=shift_query.iloc[0, 0]
        finis=shift_query.iloc[0, 1]
        shift_first=shift_query.iloc[0, 2]
        note =str(shift_query.iloc[0, 3])
        if thisHour<=finis and thisHour>=start:
            shift=shift_first
        elif thisHour>finis:
            if shift_first=='R':
                shift='B'
            else:
                shift='R'
        else:
            shift=''
    return shift, note

def get_last_shift(currShift, todayQ, yestQ, uptoHours_int, weekDate, note):
    lastShift='R'
    if currShift=='R':
        lastShift='B'
    if 'Sat' in weekDate and 'Long' in note:
        lastShift=currShift
        return lastShift, yestQ
    if 'Sun' in weekDate and 'Long' in note:
        yesterday= datetime.now()-timedelta(days=2)
        yesterday_str = yesterday.strftime("%Y-%m-%d")
        yest, yestQuery, yestQ = get_date_format(yesterday_str)
        return lastShift, yestQ
    if uptoHours_int<=14:
        return lastShift, yestQ
    else:
        return lastShift, todayQ

#=====Daily Bundle==========================
if __name__=="__main__":
    today         = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    thisWeekDate  = datetime.now().strftime('%A')
    yesterday    = datetime.now()-timedelta(days=1)
    yesterday_str = yesterday.strftime("%Y-%m-%d")
    date, dateQuery, dateQ = get_date_format(today)
    yest, yestQuery, yestQ = get_date_format(yesterday_str)
    h=today[11:13]
    uptoHours=today[11:13]
    uptoHours_int=int(uptoHours)-1
    uptoHours=str(uptoHours_int)
    today_date=today[0:10]
    if uptoHours_int<10:
        uptoHours='0'+uptoHours
    fromHours_int=uptoHours_int-1
    fromHours=str(fromHours_int)
    today_date=today[0:10]
    if fromHours_int<10:
        fromHours='0'+fromHours
    try:
    # if uptoHours_int>=6 and uptoHours_int<=24:
        dailyBundle_link='\\\\pbvfps1\\PBShare2\\Scan\\Report\\QCReconcile\\'
        thisWeek=get_this_week()
        if not os.path.exists(dailyBundle_link+thisWeek):
            os.makedirs(dailyBundle_link+thisWeek)
        if not os.path.exists(dailyBundle_link+thisWeek+'\\'+date):
            os.makedirs(dailyBundle_link+thisWeek+'\\'+date)
        currShift, note=getShiftByDate(date, int(h))
        if currShift=='':
            quit()
        lastShift, yestQ=get_last_shift(currShift, dateQ, yestQ, uptoHours_int, thisWeekDate, note)
        sql=("SELECT Ticket, scan.EMPLOYEE as ID_5, ID as ID_6, NAME, LINE, LEFT(Shift,1) as SHIFT, OPERATION, scan.WORK_LOT, active.LOCATION as GROUP_WL, SELLING_STYLE, "
            +" scan.SIZE, scan.COLOR, IRR, CONCAT(LEFT(FILE, 3),'-',MID(FILE,4,3)) AS GROUP_SCAN, MID(FILE,9, 10) as DATE_SCAN, MID(FILE, 20, 2) as HOUR_SCAN, FILE, QC "
            +" FROM (SELECT qc.TICKET, scan.EMPLOYEE, qc.IRR, scan.WORK_LOT, qc.FILE, scan.OPERATION, scan.SIZE, scan.COLOR, scan.QC FROM employee_scanticket scan "
            +" INNER JOIN qc_endline_record qc on scan.TICKET=qc.TICKET ) as scan LEFT JOIN worklot_active active ON scan.WORK_LOT=active.WORK_LOT "
            +" left join (select ID, RIGHT(ID, 5) as EMPLOYEE, Name, Line, Shift from erpsystem.setup_emplist) as Temp1 on Temp1.EMPLOYEE=scan.EMPLOYEE "
            +" WHERE (FILE like '%"+lastShift+'_'+yestQ+"%' or FILE like '%"+currShift+'_'+dateQ+"%') AND IRR!='000' order by DATE_SCAN, HOUR_SCAN, GROUP_WL;")
        if 'Long' in note and 'Sat' in thisWeekDate:
            sql=("SELECT Ticket, scan.EMPLOYEE as ID_5, ID as ID_6, NAME, LINE, LEFT(Shift,1) as SHIFT, OPERATION, scan.WORK_LOT, active.LOCATION as GROUP_WL, SELLING_STYLE, "
            +" scan.SIZE, scan.COLOR, IRR, CONCAT(LEFT(FILE, 3),'-',MID(FILE,4,3)) AS GROUP_SCAN, MID(FILE,9, 10) as DATE_SCAN, MID(FILE, 20, 2) as HOUR_SCAN, FILE, QC "
            +" FROM (SELECT qc.TICKET, scan.EMPLOYEE, qc.IRR, scan.WORK_LOT, qc.FILE, scan.OPERATION, scan.SIZE, scan.COLOR, scan.QC FROM employee_scanticket scan "
            +" INNER JOIN qc_endline_record qc on scan.TICKET=qc.TICKET ) as scan LEFT JOIN worklot_active active ON scan.WORK_LOT=active.WORK_LOT "
            +" left join (select ID, RIGHT(ID, 5) as EMPLOYEE, Name, Line, Shift from erpsystem.setup_emplist) as Temp1 on Temp1.EMPLOYEE=scan.EMPLOYEE "
            +" WHERE (FILE like '%"+currShift+'_'+dateQ+"%') AND IRR!='000' order by DATE_SCAN, HOUR_SCAN, GROUP_WL;")
        # if 'Long' in note and 'Sun' in thisWeekDate:
        #     if currShift
        #     sql=("SELECT Ticket, scan.EMPLOYEE as ID_5, ID as ID_6, NAME, LINE, LEFT(Shift,1) as SHIFT, OPERATION, scan.WORK_LOT, active.LOCATION as GROUP_WL, SELLING_STYLE, "
        #     +" scan.SIZE, scan.COLOR, IRR, CONCAT(LEFT(FILE, 3),'-',MID(FILE,4,3)) AS GROUP_SCAN, MID(FILE,9, 10) as DATE_SCAN, MID(FILE, 20, 2) as HOUR_SCAN, FILE, QC "
        #     +" FROM (SELECT qc.TICKET, scan.EMPLOYEE, qc.IRR, scan.WORK_LOT, qc.FILE, scan.OPERATION, scan.SIZE, scan.COLOR, scan.QC FROM employee_scanticket scan "
        #     +" INNER JOIN qc_endline_record qc on scan.TICKET=qc.TICKET ) as scan LEFT JOIN worklot_active active ON scan.WORK_LOT=active.WORK_LOT "
        #     +" left join (select ID, RIGHT(ID, 5) as EMPLOYEE, Name, Line, Shift from erpsystem.setup_emplist) as Temp1 on Temp1.EMPLOYEE=scan.EMPLOYEE "
        #     +" WHERE (FILE like '%"+lastShift+'_'+yestQ+"%' or FILE like '%"+currShift+'_'+dateQ+"%') AND IRR!='000' order by DATE_SCAN, HOUR_SCAN, GROUP_WL;")
        
        data=pd.read_sql(sql, engine)
        data_json=[]
        wb = openpyxl.load_workbook(r'C:\HanesApp_Report\public\Python\QC\Template\EndlineTemplate.xlsx')#dailyOutput_link+thisWeek+'\\'+date+'\\'+today_date+'_'+thisHour+'h50.xlsx')  
        row_in_shift=1
        # row_out_shift=1
        shift_real=currShift#getShift(thisWeek, uptoHours)
        for row in range(0, len(data)):
            ticket=str(data.iloc[row, 0])
            ID=data.iloc[row, 2]
            name=data.iloc[row, 3]
            line=data.iloc[row, 4]
            shift=data.iloc[row, 5]
            op=data.iloc[row, 6]
            wl=data.iloc[row, 7]
            wl_group=data.iloc[row, 8]
            selling=data.iloc[row, 9]
            size=str(data.iloc[row, 10])
            if size=='None':
                size=''
            color=str(data.iloc[row, 11])
            if color=='None':
                color=''
            irr=data.iloc[row, 12]
            scan_group=data.iloc[row, 13]
            scan_day=data.iloc[row, 14]
            scan_hour=data.iloc[row, 15]
            file=data.iloc[row, 16]
            qc=data.iloc[row, 17]
            if shift==shift_real or (shift!=shift_real and scan_day==dateQ):
                wb.active=0
                sheet = wb.active
                row_in_shift+=1
                sheet.cell(row=row_in_shift, column=1).value=ticket
                sheet.cell(row=row_in_shift, column=2).value=ID
                sheet.cell(row=row_in_shift, column=3).value=name
                sheet.cell(row=row_in_shift, column=4).value=line
                sheet.cell(row=row_in_shift, column=5).value=shift_real
                sheet.cell(row=row_in_shift, column=6).value=shift
                sheet.cell(row=row_in_shift, column=7).value=op
                sheet.cell(row=row_in_shift, column=8).value=wl
                sheet.cell(row=row_in_shift, column=9).value=wl_group
                sheet.cell(row=row_in_shift, column=10).value=selling
                sheet.cell(row=row_in_shift, column=11).value=size
                sheet.cell(row=row_in_shift, column=12).value=color
                sheet.cell(row=row_in_shift, column=13).value=irr
                sheet.cell(row=row_in_shift, column=14).value=scan_group
                sheet.cell(row=row_in_shift, column=15).value=scan_day
                sheet.cell(row=row_in_shift, column=16).value=scan_hour
                sheet.cell(row=row_in_shift, column=17).value=str(file)
                sheet.cell(row=row_in_shift, column=18).value=qc
        wb.active=1
        sheet = wb.active
        group_list=data['GROUP_SCAN'].unique()
        group_list.sort()
        for i in range(0, len(group_list)):
            group=group_list[i]
            group_query=data.query('GROUP_SCAN=="'+group+'"')
            de_lai=0
            trong_ca=0
            ngoai_ca=0
            for j in range(0, len(group_query)):
                file=group_query.iloc[j, 16]
                shift_emp =group_query.iloc[j, 5]
                shift_scan=group_query.iloc[j, 16][6:7]
                # shift_real=getShift(thisWeek, uptoHours)
                if shift_emp==shift_real and shift_scan==shift_real:
                    trong_ca+=1
                elif shift_emp!=shift_real and shift_scan==shift_real:
                    ngoai_ca+=1
                elif shift_emp==shift_real and shift_scan!=shift_real:
                    de_lai+=1
            total=trong_ca+ngoai_ca+de_lai
            sheet.cell(row=i+3, column=9).value=trong_ca
            sheet.cell(row=i+3, column=1).value=group
            sheet.cell(row=i+3, column=2).value=shift_real
            sheet.cell(row=i+3, column=3).value=de_lai
            if shift_real=='R':
                sheet.cell(row=i+3, column=4).value=trong_ca
                sheet.cell(row=i+3, column=5).value=ngoai_ca
            else:
                sheet.cell(row=i+3, column=5).value=trong_ca
                sheet.cell(row=i+3, column=4).value=ngoai_ca
            sheet.cell(row=i+3, column=6).value=total
        wb.active=2
        sheet = wb.active
        query=("select CONCAT(LEFT(FILE, 3),'-',MID(FILE,4,3)) AS GROUP_SCAN, COUNT(DISTINCT BUNDLE) as SAMPLE "
                +"from employee_scanticket where FILE like '%"+currShift+'_'+dateQ+"%' and QC!='000000' and QC!='' and QC!='999999' GROUP BY LEFT(FILE, 6);")
        data_sample=pd.read_sql(query, engine)
        if len(data_sample)>0:
            for i in range(0, len(data_sample)):
                group=str(data_sample.iloc[i, 0])
                sample=data_sample.iloc[i, 1]
                sheet.cell(row=i+2, column=1).value=group
                sheet.cell(row=i+2, column=2).value=sample
        wb.save(dailyBundle_link+thisWeek+'\\'+date+'\\QCReconcile_'+h+'.xlsx') 
        print(date+'-'+h+'h')
    except:
        print('fail;-;-;-')
        